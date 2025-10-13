from typing import Any, Dict, List
from dataclasses import dataclass
from openpyxl import load_workbook
import re

REQUIRED_MAPPING_COLUMNS = [
    "Field","Method","Number","Anchor","Label","Segment","TableIndex","Row","Col",
    "Target","Occur","Default","Transform"
]

@dataclass
class ExcelConfig:
    data_rows: List[Dict[str, Any]]
    mapping_rows: List[Dict[str, Any]]
    settings: Dict[str, Any]

def _norm_header(v) -> str:
    # сжимаем ВСЕ виды пробелов (включая неразрывный) и обрезаем
    return re.sub(r"\s+", " ", ("" if v is None else str(v))).strip()

def _norm_rule(d: dict) -> dict:
    out = {}
    for k, v in d.items():
        nk = _norm_header(k)
        if isinstance(v, str):
            out[nk] = v.strip()
        else:
            out[nk] = v
    return out

def load_excel(xlsx_path: str) -> ExcelConfig:
    wb = load_workbook(xlsx_path, data_only=True)

    # ---------- data ----------
    ws = wb["data"]
    headers = [_norm_header(c.value) for c in ws[1]]
    data_rows: List[Dict[str, Any]] = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if any(v is not None and str(v).strip() != "" for v in r):
            row = {}
            for h, v in zip(headers, r):
                row[h] = v
            data_rows.append(row)

    # ---------- mapping ----------
    ws = wb["mapping"]
    map_headers = [_norm_header(c.value) for c in ws[1]]
    missing = [h for h in REQUIRED_MAPPING_COLUMNS if h not in map_headers]
    if missing:
        raise SystemExit(4)
    mapping_rows: List[Dict[str, Any]] = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if any(v is not None and str(v).strip() != "" for v in r):
            mapping_rows.append(_norm_rule(dict(zip(map_headers, r))))

    # ---------- settings ----------
    settings: Dict[str, Any] = {}
    if "settings" in wb.sheetnames:
        ws = wb["settings"]
        for key, val in ws.iter_rows(min_row=2, max_col=2, values_only=True):
            k = _norm_header(key)
            if k:
                settings[k] = val

    return ExcelConfig(data_rows, mapping_rows, settings)
